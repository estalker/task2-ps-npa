from __future__ import annotations

import csv
import html
import hashlib
import os
import re
import subprocess
import tempfile
from collections.abc import Callable
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal
from urllib.parse import urlparse

from neo4j import Driver, GraphDatabase

from app.neo4j_stage_reset import reset_matching_layer, reset_npa_subgraph, reset_profstandard_subgraph
from fastapi import Body, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse, PlainTextResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles

KIND = Literal["ps", "npa"]
PROJ_KIND = Literal["ps", "npa", "match"]

WORKSPACE = Path(os.getenv("WORKSPACE_DIR", "/workspace")).resolve()
INPUT_DIR = (WORKSPACE / "input").resolve()
OUTPUT_DIR = (WORKSPACE / "output").resolve()

PS_DIR = INPUT_DIR / "ps"
NPA_DIR = INPUT_DIR / "npa"

RESULT_MD = OUTPUT_DIR / "list_mandatory_ps.md"

COL_REPHRASED_HINT = "переформулирован"
COL_RAW_HINT = "исходный фрагмент"

STATIC_DIR = Path(__file__).resolve().parent / "static"
INDEX_HTML_PATH = STATIC_DIR / "index.html"

app = FastAPI(title="Task2 Frontend")

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


def _read_index_html() -> str:
    if not INDEX_HTML_PATH.exists():
        return "<html><body><pre>frontend_app/static/index.html not found</pre></body></html>"
    return INDEX_HTML_PATH.read_text(encoding="utf-8", errors="replace")


def _safe_filename(name: str) -> str:
    name = name.replace("\\", "/").split("/")[-1]
    # keep basic set; prevent sneaky paths
    name = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._() \\-]+", "_", name)
    return name.strip() or "file"


def _kind_dir(kind: KIND) -> Path:
    if kind == "ps":
        return PS_DIR
    return NPA_DIR


def _ensure_dirs() -> None:
    PS_DIR.mkdir(parents=True, exist_ok=True)
    NPA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _existing_hashes(d: Path) -> dict[str, str]:
    """
    sha256 -> filename
    Best-effort: ignores unreadable files.
    """
    out: dict[str, str] = {}
    for p in d.iterdir():
        if not p.is_file():
            continue
        try:
            h = _sha256_bytes(p.read_bytes())
            out[h] = p.name
        except Exception:
            continue
    return out


def _dedup_name(d: Path, name: str) -> str:
    """
    If file exists, return a non-colliding name by adding " (2)", " (3)", ...
    """
    base = _safe_filename(name)
    stem = base
    ext = ""
    if "." in base:
        stem = base.rsplit(".", 1)[0]
        ext = "." + base.rsplit(".", 1)[1]
    cand = stem + ext
    i = 2
    while (d / cand).exists():
        cand = f"{stem} ({i}){ext}"
        i += 1
        if i > 99:
            break
    return cand


def _strip_html(s: str) -> str:
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.IGNORECASE)
    s = re.sub(r"</?small>", "", s, flags=re.IGNORECASE)
    s = re.sub(r"<[^>]+>", "", s)
    return " ".join(s.split()).strip()


def _parse_markdown_table(md_text: str) -> tuple[list[str], list[list[str]]]:
    lines = [ln.strip("\n") for ln in (md_text or "").splitlines()]
    table_lines = [ln for ln in lines if ln.strip().startswith("|")]
    if len(table_lines) < 2:
        return [], []
    header = [c.strip() for c in table_lines[0].strip("|").split("|")]
    rows: list[list[str]] = []
    for ln in table_lines[2:]:
        if not ln.strip().startswith("|"):
            continue
        cols = [c.strip() for c in ln.strip("|").split("|")]
        # pad/truncate to header length
        if len(cols) < len(header):
            cols += [""] * (len(header) - len(cols))
        rows.append(cols[: len(header)])
    return header, rows


def _rows_to_html_table(header: list[str], rows: list[list[str]], *, with_actions: bool = True) -> str:
    if not header:
        return "<div class='muted'>Нет таблицы в output/list_mandatory_ps.md</div>"
    out = ["<table>"]
    out.append("<thead><tr>")
    for h in header:
        out.append(f"<th>{html.escape(h)}</th>")
    if with_actions:
        out.append("<th>Действия</th>")
    out.append("</tr></thead>")
    out.append("<tbody>")
    for row_idx, r in enumerate(rows):
        out.append("<tr>")
        for c in r:
            # markdown output already uses <br>, <small>; keep as-is (trusted local artifact)
            out.append(f"<td>{c}</td>")
        if with_actions:
            out.append("<td>" f"<button class='secondary' onclick='rephraseRow({row_idx})'>Переделать</button>" "</td>")
        out.append("</tr>")
    out.append("</tbody></table>")
    return "".join(out)


def _rows_to_csv(header: list[str], rows: list[list[str]]) -> str:
    buf = StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(header)
    for r in rows:
        w.writerow([_strip_html(c) for c in r])
    return buf.getvalue()


def _xlsx_visual_lines(text: str, col_width: float) -> int:
    """Число визуальных строк с учётом явных \\n и переноса по ширине колонки (в усл. единицах Excel)."""
    if not text:
        return 1
    chars_per_line = max(5, int((col_width or 10) * 0.92))
    total = 0
    for chunk in str(text).split("\n"):
        chunk = chunk.strip() if chunk else ""
        if not chunk:
            total += 1
            continue
        total += max(1, (len(chunk) + chars_per_line - 1) // chars_per_line)
    return max(total, 1)


def _apply_xlsx_column_widths_and_row_heights(ws) -> None:
    from openpyxl.utils import get_column_letter

    if ws.max_row < 1 or ws.max_column < 1:
        return

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            s = str(val) if val is not None else ""
            for line in s.split("\n"):
                max_len = max(max_len, len(line))
        # Ширина колонки в Excel: ~число символов стандартного шрифта
        width = float(min(max(max_len + 2.2, 9.0), 70.0))
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            s = str(val) if val is not None else ""
            cw = float(ws.column_dimensions[get_column_letter(col)].width or 10.0)
            max_lines = max(max_lines, _xlsx_visual_lines(s, cw))
        # Высота строки в пунктах (1 pt ≈ 1/72 дюйма)
        height = min(10.0 + 13.8 * max_lines, 250.0)
        if row == 1:
            height = max(height, 26.0)
        ws.row_dimensions[row].height = height


def _rows_to_xlsx_bytes(header: list[str], rows: list[list[str]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Обязательные ПС"

    if not header:
        ws.append(["Нет таблицы в output/list_mandatory_ps.md"])
        ws.column_dimensions["A"].width = 72
        ws.row_dimensions[1].height = 22
    else:
        ws.append(header)
        for r in rows:
            ws.append([_strip_html(c) for c in r])
        bold = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical="top")
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = wrap
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(header)):
            for cell in row:
                cell.alignment = wrap
        _apply_xlsx_column_widths_and_row_heights(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _run_cmd_stream(args: list[str], cwd: Path):
    """
    Stream stdout/stderr progressively while process runs.
    We read bytes chunks (not lines) so progress bars (\\r) appear immediately.
    """
    env = os.environ.copy()
    env.setdefault("PYTHONUNBUFFERED", "1")

    p = subprocess.Popen(
        args,
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=False,
        env=env,
        bufsize=0,
    )
    assert p.stdout is not None
    try:
        while True:
            chunk = p.stdout.read(1024)
            if chunk:
                yield chunk.decode("utf-8", errors="replace")
                continue
            if p.poll() is not None:
                break
    finally:
        try:
            p.stdout.close()
        except Exception:
            pass
    rc = p.wait()
    yield f"\n[exit_code={rc}]\n"
    return rc


def _find_col_idx(header: list[str], hint: str) -> int:
    h = [str(x or "").strip().lower() for x in (header or [])]
    hint = (hint or "").strip().lower()
    for i, name in enumerate(h):
        if hint and hint in name:
            return i
    return -1


def _parse_markdown_table_with_span(md_text: str) -> tuple[list[str], list[list[str]], int, int, list[str]]:
    """
    Returns: header, rows, start_line_idx, end_line_idx_exclusive, all_lines (without \\n).
    Table span is the maximal contiguous block of lines starting with '|'.
    """
    all_lines = [ln.rstrip("\n") for ln in (md_text or "").splitlines()]
    start = -1
    end = -1
    for i, ln in enumerate(all_lines):
        if ln.strip().startswith("|"):
            start = i
            break
    if start == -1:
        return [], [], -1, -1, all_lines
    end = start
    while end < len(all_lines) and all_lines[end].strip().startswith("|"):
        end += 1
    header, rows = _parse_markdown_table("\n".join(all_lines[start:end]))
    return header, rows, start, end, all_lines


def _rows_to_markdown_table(header: list[str], rows: list[list[str]]) -> str:
    if not header:
        return ""
    sep = "|" + "|".join(["---"] * len(header)) + "|"
    out: list[str] = []
    out.append("| " + " | ".join([str(h or "").strip() for h in header]) + " |")
    out.append(sep)
    for r in rows:
        rr = [str(c or "").strip() for c in (r or [])]
        if len(rr) < len(header):
            rr += [""] * (len(header) - len(rr))
        out.append("| " + " | ".join(rr[: len(header)]) + " |")
    return "\n".join(out)


def _neo4j_driver() -> Driver:
    uri = os.getenv("NEO4J_URI", "bolt://neo4j:7687").strip() or "bolt://neo4j:7687"
    user = os.getenv("NEO4J_USER", "neo4j").strip() or "neo4j"
    password = os.getenv("NEO4J_PASSWORD", "neo4j_password").strip() or "neo4j_password"
    return GraphDatabase.driver(uri, auth=(user, password))


def _neo4j_stage_reset(reset_fn: Callable[[Driver], None]) -> None:
    drv = _neo4j_driver()
    try:
        reset_fn(drv)
    finally:
        drv.close()


def _reset_result_artifacts() -> None:
    """
    Ensure UI does not show stale result file between runs.
    """
    try:
        if RESULT_MD.exists():
            RESULT_MD.unlink()
    except Exception:
        pass


def _ollama_is_reachable() -> tuple[bool, str]:
    """
    Best-effort connectivity check for Ollama/OpenAI-compatible endpoint.
    """
    base_url = (os.getenv("OPENAI_BASE_URL", "") or "").strip()
    if not base_url:
        return False, "OPENAI_BASE_URL is empty"

    parsed = urlparse(base_url)
    host = parsed.hostname
    port = parsed.port or (443 if parsed.scheme == "https" else 80)
    if not host:
        return False, f"Cannot parse host from OPENAI_BASE_URL={base_url!r}"

    import socket

    try:
        with socket.create_connection((host, port), timeout=2.0):
            return True, "ok"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _neo4j_is_reachable() -> tuple[bool, str]:
    drv = _neo4j_driver()
    try:
        db = os.getenv("NEO4J_DATABASE", "neo4j").strip() or "neo4j"
        with drv.session(database=db) as s:
            s.run("RETURN 1 AS ok").consume()
        return True, "ok"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"
    finally:
        drv.close()


def _neo4j_profstandard_count() -> int:
    drv = _neo4j_driver()
    try:
        db = os.getenv("NEO4J_DATABASE", "neo4j").strip() or "neo4j"
        with drv.session(database=db) as s:
            rec = s.run("MATCH (d:Document {source: $src}) RETURN count(d) AS c", src="profstandard").single()
            return int(rec["c"] or 0) if rec else 0
    finally:
        drv.close()


@app.get("/api/projection")
def projection(kind: PROJ_KIND = Query("match"), limit: int = Query(400, ge=50, le=3000)) -> dict:
    """
    Flat JOIN-like projections for debugging (no graph visualization).
    Returns: {kind, header, rows}
    """
    if kind not in ("ps", "npa", "match"):
        raise HTTPException(status_code=400, detail="kind must be ps|npa|match")

    drv = _neo4j_driver()
    try:
        db = os.getenv("NEO4J_DATABASE", "neo4j").strip() or "neo4j"
        with drv.session(database=db) as s:
            if kind == "ps":
                header = ["ps_id", "ps_path", "otf_code", "otf_name", "role_name", "workscope", "involves_score", "involves_keywords"]
                q = """
                MATCH (ps:Document {source:'profstandard'})-[:HAS_OTF]->(o:OTF)
                OPTIONAL MATCH (o)-[:HAS_ROLE]->(r:Role)
                OPTIONAL MATCH (o)-[inv:INVOLVES]->(w:WorkScope)
                RETURN ps.id AS ps_id, ps.path AS ps_path,
                       o.code AS otf_code, o.name AS otf_name,
                       r.name AS role_name,
                       w.name AS workscope, inv.score AS involves_score, inv.keywords AS involves_keywords
                ORDER BY ps.updated_at DESC, o.code, r.name, w.name
                LIMIT $limit
                """
                rows = []
                for rec in s.run(q, limit=int(limit)):
                    rows.append(
                        [
                            rec.get("ps_id") or "",
                            rec.get("ps_path") or "",
                            rec.get("otf_code") or "",
                            rec.get("otf_name") or "",
                            rec.get("role_name") or "",
                            rec.get("workscope") or "",
                            "" if rec.get("involves_score") is None else str(rec.get("involves_score")),
                            ",".join([str(x) for x in (rec.get("involves_keywords") or []) if x]),
                        ]
                    )
                return {"kind": kind, "header": header, "rows": rows}

            if kind == "npa":
                header = ["npa_id", "npa_title", "norm_number", "workscope", "req_type", "req_text", "norm_text"]
                q = """
                MATCH (n:Norm)-[:MENTIONED_IN]->(d:Document {source:'npa'})
                OPTIONAL MATCH (n)-[:APPLIES_TO]->(w:WorkScope)
                OPTIONAL MATCH (n)-[:SETS_REQUIREMENT]->(r:Requirement)
                RETURN d.id AS npa_id, d.title AS npa_title,
                       n.number AS norm_number, w.name AS workscope,
                       r.type AS req_type, r.text AS req_text,
                       n.text AS norm_text
                ORDER BY d.updated_at DESC, n.number
                LIMIT $limit
                """
                rows = []
                for rec in s.run(q, limit=int(limit)):
                    rows.append(
                        [
                            rec.get("npa_id") or "",
                            rec.get("npa_title") or "",
                            rec.get("norm_number") or "",
                            rec.get("workscope") or "",
                            rec.get("req_type") or "",
                            (rec.get("req_text") or "")[:400],
                            (rec.get("norm_text") or "")[:500],
                        ]
                    )
                return {"kind": kind, "header": header, "rows": rows}

            # match
            header = [
                "npa_title",
                "norm_number",
                "workscope",
                "ps_id",
                "otf_code",
                "otf_name",
                "match_via_workscope",
                "involves_score",
                "involves_keywords",
            ]
            q = """
            MATCH (n:Norm)-[m:MATCHES_OTF]->(o:OTF)
            OPTIONAL MATCH (n)-[:MENTIONED_IN]->(nd:Document {source:'npa'})
            OPTIONAL MATCH (o)<-[:HAS_OTF]-(pd:Document {source:'profstandard'})
            OPTIONAL MATCH (n)-[:APPLIES_TO]->(w:WorkScope)
            OPTIONAL MATCH (o)-[inv:INVOLVES]->(w2:WorkScope {name: coalesce(w.name, m.via_workscope)})
            RETURN
              nd.title AS npa_title,
              n.number AS norm_number,
              coalesce(w.name, m.via_workscope) AS workscope,
              pd.id AS ps_id,
              o.code AS otf_code,
              o.name AS otf_name,
              m.via_workscope AS match_via_workscope,
              inv.score AS involves_score,
              inv.keywords AS involves_keywords
            ORDER BY n.number, o.code
            LIMIT $limit
            """
            rows = []
            for rec in s.run(q, limit=int(limit)):
                rows.append(
                    [
                        rec.get("npa_title") or "",
                        rec.get("norm_number") or "",
                        rec.get("workscope") or "",
                        rec.get("ps_id") or "",
                        rec.get("otf_code") or "",
                        rec.get("otf_name") or "",
                        rec.get("match_via_workscope") or "",
                        "" if rec.get("involves_score") is None else str(rec.get("involves_score")),
                        ",".join([str(x) for x in (rec.get("involves_keywords") or []) if x]),
                    ]
                )
            return {"kind": kind, "header": header, "rows": rows}
    finally:
        drv.close()


def _stream_file_bytes(path: Path, *, chunk_size: int = 1024 * 1024):
    with path.open("rb") as f:
        while True:
            b = f.read(chunk_size)
            if not b:
                break
            yield b


def _projection_query(kind: PROJ_KIND) -> tuple[list[str], str, Callable[[dict], list[str]]]:
    """
    Shared query definitions for projections.
    Note: for NPA projection we keep the same cell truncation as the UI (req_text/norm_text).
    """
    if kind == "ps":
        header = ["ps_id", "ps_path", "otf_code", "otf_name", "role_name", "workscope", "involves_score", "involves_keywords"]
        q = """
        MATCH (ps:Document {source:'profstandard'})-[:HAS_OTF]->(o:OTF)
        OPTIONAL MATCH (o)-[:HAS_ROLE]->(r:Role)
        OPTIONAL MATCH (o)-[inv:INVOLVES]->(w:WorkScope)
        RETURN ps.id AS ps_id, ps.path AS ps_path,
               o.code AS otf_code, o.name AS otf_name,
               r.name AS role_name,
               w.name AS workscope, inv.score AS involves_score, inv.keywords AS involves_keywords
        ORDER BY ps.updated_at DESC, o.code, r.name, w.name
        """

        def row(rec: dict) -> list[str]:
            return [
                rec.get("ps_id") or "",
                rec.get("ps_path") or "",
                rec.get("otf_code") or "",
                rec.get("otf_name") or "",
                rec.get("role_name") or "",
                rec.get("workscope") or "",
                "" if rec.get("involves_score") is None else str(rec.get("involves_score")),
                ",".join([str(x) for x in (rec.get("involves_keywords") or []) if x]),
            ]

        return header, q, row

    if kind == "npa":
        header = ["npa_id", "npa_title", "norm_number", "workscope", "req_type", "req_text", "norm_text"]
        q = """
        MATCH (n:Norm)-[:MENTIONED_IN]->(d:Document {source:'npa'})
        OPTIONAL MATCH (n)-[:APPLIES_TO]->(w:WorkScope)
        OPTIONAL MATCH (n)-[:SETS_REQUIREMENT]->(r:Requirement)
        RETURN d.id AS npa_id, d.title AS npa_title,
               n.number AS norm_number, w.name AS workscope,
               r.type AS req_type, r.text AS req_text,
               n.text AS norm_text
        ORDER BY d.updated_at DESC, n.number
        """

        def row(rec: dict) -> list[str]:
            return [
                rec.get("npa_id") or "",
                rec.get("npa_title") or "",
                rec.get("norm_number") or "",
                rec.get("workscope") or "",
                rec.get("req_type") or "",
                (rec.get("req_text") or "")[:400],
                (rec.get("norm_text") or "")[:500],
            ]

        return header, q, row

    # match
    header = [
        "npa_title",
        "norm_number",
        "workscope",
        "ps_id",
        "otf_code",
        "otf_name",
        "match_via_workscope",
        "involves_score",
        "involves_keywords",
    ]
    q = """
    MATCH (n:Norm)-[m:MATCHES_OTF]->(o:OTF)
    OPTIONAL MATCH (n)-[:MENTIONED_IN]->(nd:Document {source:'npa'})
    OPTIONAL MATCH (o)<-[:HAS_OTF]-(pd:Document {source:'profstandard'})
    OPTIONAL MATCH (n)-[:APPLIES_TO]->(w:WorkScope)
    OPTIONAL MATCH (o)-[inv:INVOLVES]->(w2:WorkScope {name: coalesce(w.name, m.via_workscope)})
    RETURN
      nd.title AS npa_title,
      n.number AS norm_number,
      coalesce(w.name, m.via_workscope) AS workscope,
      pd.id AS ps_id,
      o.code AS otf_code,
      o.name AS otf_name,
      m.via_workscope AS match_via_workscope,
      inv.score AS involves_score,
      inv.keywords AS involves_keywords
    ORDER BY n.number, o.code
    """

    def row(rec: dict) -> list[str]:
        return [
            rec.get("npa_title") or "",
            rec.get("norm_number") or "",
            rec.get("workscope") or "",
            rec.get("ps_id") or "",
            rec.get("otf_code") or "",
            rec.get("otf_name") or "",
            rec.get("match_via_workscope") or "",
            "" if rec.get("involves_score") is None else str(rec.get("involves_score")),
            ",".join([str(x) for x in (rec.get("involves_keywords") or []) if x]),
        ]

    return header, q, row


def _llm_subpayload(payload: dict, key: str) -> dict:
    b = payload.get(key)
    return b if isinstance(b, dict) else {}


def _resolve_stage_llm(payload: dict, key: str) -> tuple[bool, str | None]:
    """
    use_llm + model for one of: ps | npa | table.
    Legacy fields use_llm / llm_model apply when stage is 'all' and nested keys omit use_llm.
    """
    stage = (str(payload.get("stage") or "all")).strip() or "all"
    leg_u = bool(payload.get("use_llm", False))
    lm = payload.get("llm_model", None)
    if lm is not None and not isinstance(lm, str):
        lm = None
    lm = (lm or "").strip() or None

    block = _llm_subpayload(payload, key)
    u = block.get("use_llm")
    if u is None:
        u = leg_u if stage == "all" else False
    else:
        u = bool(u)

    m = block.get("model")
    if m is not None and not isinstance(m, str):
        m = None
    m = (m or "").strip() or None
    if m is None and stage == "all":
        m = lm
    return u, m


def _resolve_stage_reset(payload: dict, key: str, *, default: bool = True) -> bool:
    """
    reset_db flag for ps|npa|table. Default keeps legacy behavior (reset on).
    """
    block = _llm_subpayload(payload, key)
    v = block.get("reset_db")
    if v is None:
        return default
    return bool(v)


def _snapshot_openai_model() -> str | None:
    v = os.environ.get("OPENAI_MODEL")
    return v if isinstance(v, str) and v.strip() else None


def _restore_openai_model(prev: str | None) -> None:
    if prev is None:
        os.environ.pop("OPENAI_MODEL", None)
    else:
        os.environ["OPENAI_MODEL"] = prev


def _apply_openai_model_for_llm(use: bool, model: str | None) -> str | None:
    prev = _snapshot_openai_model()
    if use and model:
        os.environ["OPENAI_MODEL"] = model
    return prev


def _pipeline_chunks(payload: dict):
    _ensure_dirs()
    stage = (str(payload.get("stage") or "all")).strip() or "all"
    if stage not in ("ps", "npa", "table", "all"):
        yield f"[ERROR] Неизвестный stage={stage!r} (ожидается ps|npa|table|all)\n"
        return

    use_ps, model_ps = _resolve_stage_llm(payload, "ps")
    use_npa, model_npa = _resolve_stage_llm(payload, "npa")
    use_tbl, model_tbl = _resolve_stage_llm(payload, "table")
    reset_ps = _resolve_stage_reset(payload, "ps", default=True)
    reset_npa = _resolve_stage_reset(payload, "npa", default=True)

    want_llm = False
    if stage in ("ps", "all") and use_ps:
        want_llm = True
    if stage in ("npa", "all") and use_npa:
        want_llm = True
    if stage in ("table", "all") and use_tbl:
        want_llm = True

    if want_llm:
        ok, why = _ollama_is_reachable()
        if not ok:
            yield f"[LLM] Запрошена нейросеть, но сервер недоступен ({why}); LLM-шаги будут пропущены.\n\n"
            use_ps = use_npa = use_tbl = False

    yield (
        f"[LLM] stage={stage} "
        f"ps={use_ps} model={model_ps or '(по умолчанию)'} | "
        f"npa={use_npa} model={model_npa or '(по умолчанию)'} | "
        f"table={use_tbl} model={model_tbl or '(по умолчанию)'}\n\n"
    )

    if stage in ("ps", "all"):
        if reset_ps:
            try:
                _neo4j_stage_reset(reset_profstandard_subgraph)
                _reset_result_artifacts()
                yield "[NEO4J] сброс этапа 1: только ПС (profstandard, OTF/Role, связанные Profession/Qualification/Requirement)\n\n"
                yield "[OUTPUT] cleared output/list_mandatory_ps.md\n\n"
            except Exception as e:
                yield f"[ERROR] Neo4j reset (этап 1) failed: {e}\n"
                return
        else:
            yield "[NEO4J] этап 1: сброс отключён (добавляем/обновляем ПС поверх имеющихся данных)\n\n"

        ps_files = [p for p in PS_DIR.iterdir() if p.is_file()]
        if not ps_files:
            yield "[ERROR] Нет файлов ПС в input/ps (нужно загрузить .docx)\n"
            return
        if not any(p.suffix.lower() == ".docx" for p in ps_files):
            yield "[ERROR] В input/ps нет .docx (текущий ingest ПС работает с .docx)\n"
            return

        prev_m = _apply_openai_model_for_llm(use_ps, model_ps)
        rc = 0
        try:
            cmd = ["python", "-u", "-m", "app.ingest", "--input", "input/ps", "--doc-source", "profstandard"]
            if not use_ps:
                cmd.append("--no-llm")
            else:
                cmd += ["--llm-max-chunks", "4"]
            yield "[CMD] " + " ".join(cmd) + "\n"
            for chunk in _run_cmd_stream(cmd, cwd=WORKSPACE):
                yield chunk
                if chunk.startswith("\n[exit_code="):
                    try:
                        rc = int(chunk.split("=", 1)[1].split("]", 1)[0])
                    except Exception:
                        rc = 1
        finally:
            _restore_openai_model(prev_m)

        if rc != 0:
            yield "[ERROR] PS ingest failed\n"
            return

        if stage == "ps":
            yield "\n[OK] этап 1 (ПС → Neo4j) завершён\n"
            return

    if stage in ("npa", "all"):
        if stage == "npa" and _neo4j_profstandard_count() == 0:
            yield "[ERROR] В Neo4j нет ПС (profstandard). Сначала выполните этап 1.\n"
            return

        if reset_npa:
            try:
                _neo4j_stage_reset(reset_npa_subgraph)
                _reset_result_artifacts()
                yield "[NEO4J] сброс этапа 2: только НПА (Document npa, Norm, осиротевшие WorkScope/Requirement)\n\n"
                yield "[OUTPUT] cleared output/list_mandatory_ps.md\n\n"
            except Exception as e:
                yield f"[ERROR] Neo4j reset (этап 2) failed: {e}\n"
                return
        else:
            yield "[NEO4J] этап 2: сброс отключён (добавляем/обновляем НПА поверх имеющихся данных)\n\n"

        npa_files = [p for p in NPA_DIR.iterdir() if p.is_file()]
        if any(p.suffix.lower() == ".rtf" for p in npa_files):
            prev_m = _apply_openai_model_for_llm(use_npa, model_npa)
            rc = 0
            try:
                cmd = ["python", "-u", "-m", "app.npa_ingest", "--input", "input/npa"]
                if use_npa:
                    cmd += ["--use-llm", "--llm-max-calls", "6"]
                yield "\n[CMD] " + " ".join(cmd) + "\n"
                for chunk in _run_cmd_stream(cmd, cwd=WORKSPACE):
                    yield chunk
                    if chunk.startswith("\n[exit_code="):
                        try:
                            rc = int(chunk.split("=", 1)[1].split("]", 1)[0])
                        except Exception:
                            rc = 1
            finally:
                _restore_openai_model(prev_m)

            if rc != 0:
                yield "[ERROR] NPA ingest failed\n"
                return
        else:
            yield "\n[NPA] Нет .rtf в input/npa; app.npa_ingest пропущен.\n"

        if stage == "npa":
            yield "\n[OK] этап 2 (НПА → Neo4j) завершён\n"
            return

    if stage in ("table", "all"):
        try:
            _neo4j_stage_reset(reset_matching_layer)
            yield "[NEO4J] сброс этапа 3: слой сопоставления (OTF, Role и их связи)\n\n"
        except Exception as e:
            yield f"[ERROR] Neo4j reset (этап 3) failed: {e}\n"
            return

        _reset_result_artifacts()
        yield "[OUTPUT] cleared output/list_mandatory_ps.md (перед экспортом)\n\n"

        cmd = ["python", "-u", "scripts/build_matching_graph.py"]
        yield "\n[CMD] " + " ".join(cmd) + "\n"
        rc = 0
        for chunk in _run_cmd_stream(cmd, cwd=WORKSPACE):
            yield chunk
            if chunk.startswith("\n[exit_code="):
                try:
                    rc = int(chunk.split("=", 1)[1].split("]", 1)[0])
                except Exception:
                    rc = 1
        if rc != 0:
            yield "[ERROR] Matching graph build failed\n"
            return

        prev_m = _apply_openai_model_for_llm(use_tbl, model_tbl)
        try:
            cmd = ["python", "-u", "scripts/export_mandatory_ps_table.py"]
            if use_tbl:
                cmd += ["--rephrase-with-llm", "--llm-max-calls", "50"]
            yield "\n[CMD] " + " ".join(cmd) + "\n"
            rc = 0
            for chunk in _run_cmd_stream(cmd, cwd=WORKSPACE):
                yield chunk
                if chunk.startswith("\n[exit_code="):
                    try:
                        rc = int(chunk.split("=", 1)[1].split("]", 1)[0])
                    except Exception:
                        rc = 1
        finally:
            _restore_openai_model(prev_m)

        if rc != 0:
            yield "[ERROR] Export table failed\n"
            return

    yield "\n[OK] done\n"


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return _read_index_html()


@app.get("/api/health")
def health() -> dict:
    neo_ok, neo_detail = _neo4j_is_reachable()
    llm_ok, llm_detail = _ollama_is_reachable()
    return {
        "frontend": {"ok": True, "detail": "ok"},
        "neo4j": {"ok": neo_ok, "detail": neo_detail},
        "ollama": {"ok": llm_ok, "detail": llm_detail},
    }


@app.get("/api/files")
def list_files(kind: KIND = Query(...)) -> dict:
    _ensure_dirs()
    d = _kind_dir(kind)
    files = sorted([p.name for p in d.iterdir() if p.is_file()])
    return {"kind": kind, "files": files}


@app.post("/api/upload")
async def upload_files(kind: KIND = Query(...), files: list[UploadFile] = File(...)) -> dict:
    _ensure_dirs()
    d = _kind_dir(kind)
    saved: list[str] = []
    skipped: list[str] = []
    hashes = _existing_hashes(d)
    for f in files:
        name = _safe_filename(f.filename or "file")
        if not name.lower().endswith((".docx", ".rtf")):
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {name}")
        body = await f.read()
        h = _sha256_bytes(body)
        if h in hashes:
            skipped.append(name)
            continue
        dst_name = name if not (d / name).exists() else _dedup_name(d, name)
        dst = d / dst_name
        with dst.open("wb") as w:
            w.write(body)
        hashes[h] = dst_name
        saved.append(dst_name)
    return {"saved": saved, "skipped_duplicates": skipped}


@app.delete("/api/files/{kind}/{name}")
def delete_file(kind: KIND, name: str) -> dict:
    _ensure_dirs()
    d = _kind_dir(kind)
    safe = _safe_filename(name)
    p = d / safe
    if p.exists() and p.is_file():
        p.unlink()
    return {"deleted": safe}


@app.post("/api/run")
def run_pipeline(payload: dict = Body(default_factory=dict)) -> dict:
    parts: list[str] = []
    for ch in _pipeline_chunks(payload):
        parts.append(ch)
    text = "".join(parts)
    if "[ERROR]" in text:
        raise HTTPException(status_code=500, detail=text[-4000:])
    return {"ok": True, "log": text}


@app.post("/api/run.stream")
def run_pipeline_stream(payload: dict = Body(default_factory=dict)):
    return StreamingResponse(
        _pipeline_chunks(payload),
        media_type="text/plain; charset=utf-8",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/result.md", response_class=PlainTextResponse)
def result_md() -> str:
    if not RESULT_MD.exists():
        return ""
    return RESULT_MD.read_text(encoding="utf-8", errors="replace")


@app.get("/api/result.html", response_class=HTMLResponse)
def result_html() -> str:
    if not RESULT_MD.exists():
        return "<div class='muted'>Файл output/list_mandatory_ps.md пока не создан.</div>"
    md = RESULT_MD.read_text(encoding="utf-8", errors="replace")
    header, rows = _parse_markdown_table(md)
    return _rows_to_html_table(header, rows, with_actions=True)


@app.post("/api/rephrase-row", response_class=HTMLResponse)
def rephrase_row(row: int = Query(..., ge=0)) -> str:
    if not RESULT_MD.exists():
        raise HTTPException(status_code=404, detail="output/list_mandatory_ps.md not found")

    md = RESULT_MD.read_text(encoding="utf-8", errors="replace")
    header, rows, start, end, all_lines = _parse_markdown_table_with_span(md)
    if not header or not rows or start < 0 or end < 0:
        raise HTTPException(status_code=400, detail="No markdown table found")
    if row < 0 or row >= len(rows):
        raise HTTPException(status_code=400, detail=f"Row index out of range: {row} (0..{len(rows)-1})")

    i_rephr = _find_col_idx(header, COL_REPHRASED_HINT)
    i_raw = _find_col_idx(header, COL_RAW_HINT)
    if i_rephr < 0 or i_raw < 0:
        raise HTTPException(status_code=400, detail="Cannot locate required columns in header")

    raw_cell = rows[row][i_raw] if i_raw < len(rows[row]) else ""
    raw_text = _strip_html(raw_cell)
    if not raw_text or raw_text == "—":
        raise HTTPException(status_code=400, detail="Raw snippet is empty for this row")

    from app.table_llm_rephrase import try_rephrase_table_snippet

    out = try_rephrase_table_snippet(raw_text)
    if not out:
        raise HTTPException(status_code=500, detail="LLM rephrase returned empty result (check OPENAI_BASE_URL / model)")

    rows[row][i_rephr] = html.escape(out)

    new_tbl = _rows_to_markdown_table(header, rows)
    if not new_tbl.strip():
        raise HTTPException(status_code=500, detail="Failed to rebuild markdown table")

    new_lines = list(all_lines)
    new_lines[start:end] = new_tbl.splitlines()
    RESULT_MD.write_text("\n".join(new_lines) + "\n", encoding="utf-8")

    md2 = RESULT_MD.read_text(encoding="utf-8", errors="replace")
    header2, rows2 = _parse_markdown_table(md2)
    return _rows_to_html_table(header2, rows2, with_actions=True)


@app.get("/api/result.csv")
def result_csv() -> Response:
    if not RESULT_MD.exists():
        raise HTTPException(status_code=404, detail="output/list_mandatory_ps.md not found")
    md = RESULT_MD.read_text(encoding="utf-8", errors="replace")
    header, rows = _parse_markdown_table(md)
    body = _rows_to_csv(header, rows)
    return Response(
        content=body,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=list_mandatory_ps.csv"},
    )


@app.get("/api/result.xlsx")
def result_xlsx() -> Response:
    if not RESULT_MD.exists():
        raise HTTPException(status_code=404, detail="output/list_mandatory_ps.md not found")
    md = RESULT_MD.read_text(encoding="utf-8", errors="replace")
    header, rows = _parse_markdown_table(md)
    body = _rows_to_xlsx_bytes(header, rows)
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=list_mandatory_ps.xlsx"},
    )


@app.get("/api/projections.xlsx")
def projections_xlsx() -> StreamingResponse:
    """
    Export graph projections into a single .xlsx with 3 sheets: ПС, НПА, MATCH.
    No row limit (except Excel's own maximum).
    """
    from openpyxl import Workbook

    _ensure_dirs()
    drv = _neo4j_driver()
    tmp_path: Path | None = None
    try:
        wb = Workbook(write_only=True)
        try:
            if wb.worksheets:
                wb.remove(wb.worksheets[0])
        except Exception:
            pass

        db = os.getenv("NEO4J_DATABASE", "neo4j").strip() or "neo4j"
        with drv.session(database=db) as s:
            for kind, title in [("ps", "ПС"), ("npa", "НПА"), ("match", "MATCH")]:
                header, q, row_fn = _projection_query(kind)  # type: ignore[arg-type]
                ws = wb.create_sheet(title=title)
                ws.append(header)
                for rec in s.run(q):
                    ws.append(row_fn(rec))

        fd, tmp_name = tempfile.mkstemp(prefix="projections_", suffix=".xlsx", dir=str(OUTPUT_DIR))
        os.close(fd)
        tmp_path = Path(tmp_name)
        wb.save(tmp_path)

        def _body():
            assert tmp_path is not None
            try:
                yield from _stream_file_bytes(tmp_path)
            finally:
                try:
                    tmp_path.unlink(missing_ok=True)
                except Exception:
                    pass

        return StreamingResponse(
            _body(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=projections.xlsx"},
        )
    finally:
        drv.close()

